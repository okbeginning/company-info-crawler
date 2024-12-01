import os
import re
import time
import random
import requests
import openpyxl
from enum import Enum
from datetime import datetime
from openpyxl import styles

class ReportType(Enum):
    """报告类型枚举"""
    ANNUAL = ("年度报告", "年度报告", "001")
    SEMI_ANNUAL = ("半年度报告", "半年度报告", "002")
    Q1 = ("第一季度报告", "第一季度报告", "003")
    Q3 = ("第三季度报告", "第三季度报告", "004")
    SOCIAL = ("社会责任报告", "社会责任报告", "701")
    INTERNAL = ("内部控制报告", "内部控制报告", "702")
    ESG = ("ESG报告", "ESG报告", "703")
    SUSTAINABLE = ("可持续发展报告", "可持续发展报告", "704")
    IPO_PROSPECTUS = ("招股说明书", "招股说明书", "005")
    IPO_INQUIRY = ("问询函", "问询函回复", "006")
    
    def __init__(self, code_name, report_name, code):
        self.code_name = code_name
        self.report_name = report_name
        self.code = code
    
    @classmethod
    def from_name(cls, name):
        """根据报告名称获取报告类型"""
        for report_type in cls:
            if report_type.report_name == name:
                return report_type
        return None

class StockCrawler:
    """股票爬虫类"""
    def __init__(self, stock_code, update_progress=None):
        """
        初始化爬虫
        
        Args:
            stock_code: 股票代码
            update_progress: 更新进度的回调函数
        """
        self.stock_code = stock_code
        self.update_progress = update_progress or print
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }
        self.available_files = []  # 添加这一行
        
        # 设置下载目录
        base_dir = "financial_reports"
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        self.download_dir = os.path.join(base_dir, f"download_{timestamp}")
        os.makedirs(self.download_dir, exist_ok=True)
        
    def get_report_list(self, start_date, end_date, report_type):
        """获取指定时间范围和类型的报告列表"""
        page_index = 1
        page_size = 50
        all_reports = []
        max_retries = 3  # 最大重试次数
        base_delay = 2  # 基础延迟时间（秒）
        
        # 根据报告类型设置不同的category
        category_map = {
            ReportType.ANNUAL: "category_ndbg_szsh",
            ReportType.SEMI_ANNUAL: "category_bndbg_szsh",
            ReportType.Q1: "category_yjdbg_szsh",
            ReportType.Q3: "category_sjdbg_szsh",
            ReportType.SOCIAL: "category_shzr_szsh",
            ReportType.INTERNAL: "category_ndbg_szsh",  # 内部控制报告通常和年报一起
            ReportType.ESG: "category_shzr_szsh",  # ESG报告通常在社会责任报告分类下
            ReportType.SUSTAINABLE: "category_shzr_szsh",  # 可持续发展报告通常在社会责任报告分类下
            ReportType.IPO_PROSPECTUS: "category_zfyxs_szsh",  # 招股意向书
            ReportType.IPO_INQUIRY: "category_qita_szsh",  # 问询函通常在其他分类下
        }
        
        while True:
            params = {
                'sr': '-1',  # 按时间倒序
                'page_size': str(page_size),
                'page_index': str(page_index),
                'ann_type': 'A',  # A股
                'client_source': 'web',
                'stock_list': self.stock_code,
                'f_node': '0',
                's_node': report_type.code,
                'begin_time': start_date.strftime('%Y-%m-%d'),
                'end_time': end_date.strftime('%Y-%m-%d'),
            }
            
            # 获取对应的category
            category = category_map.get(report_type)
            if category:
                params['category'] = category
            else:
                # 如果找不到对应的category，使用所有可能的category
                params['category'] = ';'.join([
                    'category_ndbg_szsh',   # 年度报告
                    'category_bndbg_szsh',  # 半年度报告
                    'category_yjdbg_szsh',  # 一季报
                    'category_sjdbg_szsh',  # 三季报
                    'category_shzr_szsh',   # 社会责任报告
                    'category_zfyxs_szsh',  # 招股意向书
                    'category_qita_szsh'    # 其他
                ])
            
            url = 'https://np-anotice-stock.eastmoney.com/api/security/ann'
            request_url = f"{url}?{'&'.join(f'{k}={v}' for k, v in params.items())}"
            
            for retry in range(max_retries):
                try:
                    self.update_progress(f"正在请求第 {page_index} 页数据 (尝试 {retry + 1}/{max_retries})...")
                    self.update_progress(f"请求URL: {request_url}")
                    
                    # 使用递增的超时时间
                    timeout = 10 * (retry + 1)
                    response = requests.get(url, params=params, headers=self.headers, timeout=timeout)
                    self.update_progress(f"服务器响应状态码: {response.status_code}")
                    
                    # 打印原始响应数据用于调试
                    self.update_progress(f"服务器响应数据: {response.text[:500]}")  # 只打印前500个字符避免日志过长
                    
                    response.raise_for_status()
                    data = response.json()
                    
                    if not data:
                        self.update_progress("API返回数据为空")
                        return all_reports
                        
                    if 'data' not in data:
                        self.update_progress(f"API返回数据格式异常: {data}")
                        return all_reports
                        
                    if 'list' not in data['data']:
                        self.update_progress("API返回数据中没有list字段")
                        return all_reports
                        
                    reports = data['data']['list']
                    if not reports:
                        self.update_progress("没有找到更多报告")
                        return all_reports
                        
                    # 打印第一份报告的详细信息用于调试
                    if reports and page_index == 1 and retry == 0:
                        first_report = reports[0]
                        self.update_progress("第一份报告详细信息:")
                        for key, value in first_report.items():
                            self.update_progress(f"  {key}: {value}")
                    
                    # 打印所有报告的 art_code
                    for report in reports:
                        self.update_progress(f"报告标题: {report.get('title', '未知')}")
                        self.update_progress(f"  art_code: {report.get('art_code', '未知')}")
                        self.update_progress(f"  attachPath: {report.get('attachPath', '未知')}")
                        self.update_progress(f"  公告编号: {report.get('bulletin_id', '未知')}")
                        self.update_progress(f"  下载地址: https://pdf.dfcfw.com/pdf/H2_{report.get('art_code', '')}_1.pdf")
                        self.update_progress("---")
                    
                    all_reports.extend(reports)
                    self.update_progress(f"找到 {len(reports)} 份{report_type.report_name}")
                    
                    # 如果返回的数据少于page_size，说明已经是最后一页
                    if len(reports) < page_size:
                        return all_reports
                        
                    page_index += 1
                    # 添加短暂延迟避免请求过快
                    time.sleep(random.uniform(0.5, 1.5))
                    break  # 请求成功，跳出重试循环
                    
                except requests.exceptions.Timeout:
                    if retry < max_retries - 1:
                        delay = base_delay * (retry + 1)  # 使用指数退避
                        self.update_progress(f"请求超时，{delay}秒后进行第{retry + 2}次重试...")
                        time.sleep(delay)
                    else:
                        self.update_progress(f"请求超时，已达到最大重试次数")
                        return all_reports
                        
                except requests.exceptions.RequestException as e:
                    if retry < max_retries - 1:
                        delay = base_delay * (retry + 1)
                        self.update_progress(f"网络请求错误: {str(e)}")
                        self.update_progress(f"{delay}秒后进行第{retry + 2}次重试...")
                        time.sleep(delay)
                    else:
                        self.update_progress(f"网络请求错误，已达到最大重试次数: {str(e)}")
                        return all_reports
                        
                except Exception as e:
                    self.update_progress(f"获取报告列表时出错: {str(e)}")
                    return all_reports
            
        return all_reports
        
    def get_available_files(self, years=None, selected_types=None, stock_name=None):
        """
        获取可下载的文件列表
        
        Args:
            years: 要获取的年份列表，如果为None则获取近三年的报告
            selected_types: 选择的报告类型列表，如果为None则获取所有类型的报告
            stock_name: 股票名称，用于日志显示
            
        Returns:
            list: 可下载文件列表，每个文件包含标题、日期、类型等信息
        """
        # 如果未指定年份，默认获取近三年的报告
        if years is None:
            current_year = datetime.now().year
            years = list(range(current_year - 2, current_year + 1))
            
        if years:
            years.sort()  # 确保年份是有序的
            start_date = datetime(years[0], 1, 1)
            end_date = datetime(years[-1], 12, 31)
        else:
            start_date = datetime(datetime.now().year - 2, 1, 1)
            end_date = datetime(datetime.now().year, 12, 31)
            
        # 如果未指定报告类型，默认获取所有类型的报告
        if selected_types is None:
            selected_types = [t.report_name for t in ReportType]
            
        available_files = []
        for type_name in selected_types:
            report_type = ReportType.from_name(type_name)
            if report_type is None:
                self.update_progress(f"未知的报告类型: {type_name}")
                continue
                
            self.update_progress(f"正在获取{type_name}列表...")
            report_list = self.get_report_list(start_date, end_date, report_type)
            
            # 获取当前报告类型的关键词列表
            keywords = {
                ReportType.ANNUAL: ["年度报告", "年报"],
                ReportType.SEMI_ANNUAL: ["半年度报告", "半年报"],
                ReportType.Q1: ["第一季度报告", "一季报", "第1季度报告", "2024年第一季度业绩预告"],
                ReportType.Q3: ["第三季度报告", "三季报", "第3季度报告"],
                ReportType.SOCIAL: ["社会责任报告", "企业社会责任报告"],
                ReportType.INTERNAL: ["内部控制报告", "内控报告", "内控自我评价报告"],
                ReportType.ESG: ["ESG报告", "环境、社会及管治报告", "环境社会治理报告"],
                ReportType.SUSTAINABLE: ["可持续发展报告"],
                ReportType.IPO_PROSPECTUS: ["招股说明书", "招股意向书", "发行保荐书"],
                ReportType.IPO_INQUIRY: ["问询函", "回复", "审核问询函"]
            }.get(report_type, [])
            
            filtered_count = 0
            total_count = len(report_list)
            
            for report in report_list:
                try:
                    title = report['title']
                    notice_date = report['notice_date']
                    date = datetime.strptime(notice_date, '%Y-%m-%d %H:%M:%S')
                    
                    # 检查年份（如果不是IPO相关报告）
                    is_ipo_report = report_type in [ReportType.IPO_PROSPECTUS, ReportType.IPO_INQUIRY]
                    if not is_ipo_report and years and date.year not in years:
                        filtered_count += 1
                        continue
                        
                    # 检查报告类型（放宽匹配条件）
                    matched = False
                    if keywords:
                        # 1. 检查完整关键词匹配
                        if any(keyword in title for keyword in keywords):
                            matched = True
                        # 2. 对于问询函和回复，使用更宽松的匹配
                        elif report_type == ReportType.IPO_INQUIRY:
                            matched = "问询" in title or "回复" in title
                        # 3. 对于年报，检查年份+报告的组合
                        elif report_type == ReportType.ANNUAL and str(date.year) in title and "报告" in title:
                            matched = True
                        # 4. 对于季报，检查季度+报告的组合
                        elif report_type in [ReportType.Q1, ReportType.Q3]:
                            quarter_keywords = ["一季", "1季", "三季", "3季"] if report_type == ReportType.Q1 else ["三季", "3季"]
                            matched = any(qk in title and "报告" in title for qk in quarter_keywords)
                    else:
                        matched = True  # 如果没有关键词，则默认匹配
                        
                    if not matched:
                        filtered_count += 1
                        continue
                        
                    # 获取文件大小（以MB为单位）
                    file_size = report.get('file_size', 0)
                    if file_size:
                        size_str = f"{file_size / 1024 / 1024:.2f}MB"
                    else:
                        size_str = "未知"
                        
                    available_files.append({
                        'title': title,
                        'date': date,
                        'type': type_name,
                        'size': size_str,
                        'art_code': report['art_code'],
                        'download_url': f"https://pdf.dfcfw.com/pdf/H2_{report['art_code']}_1.pdf"  # 修改下载链接格式
                    })
                    
                except (ValueError, TypeError, KeyError) as e:
                    self.update_progress(f"处理报告数据时出错: {str(e)}")
                    continue
            
            if filtered_count > 0:
                self.update_progress(f"在{total_count}份文件中过滤掉{filtered_count}份不符合条件的文件")
                
        self.update_progress(f"共找到 {len(available_files)} 个可下载文件")
        self.available_files = available_files  # 添加这一行
        return available_files

    def download_file(self, file_info):
        """下载单个文件"""
        try:
            if not file_info.get('art_code'):
                self.update_progress(f"错误：无法获取文件的 art_code: {file_info}")
                return False

            # 构建下载链接
            download_url = f"https://pdf.dfcfw.com/pdf/H2_{file_info['art_code']}_1.pdf"
            self.update_progress(f"尝试下载文件: {file_info['title']}")
            self.update_progress(f"下载链接: {download_url}")

            # 创建下载目录
            os.makedirs(self.download_dir, exist_ok=True)
            
            # 构建文件名
            filename = f"{file_info['title']}_{file_info['date']}.pdf"
            filename = re.sub(r'[<>:"/\\|?*]', '_', filename)  # 替换非法字符
            filepath = os.path.join(self.download_dir, filename)
            
            # 使用 stream 方式下载文件
            response = requests.get(download_url, headers=self.headers, stream=True)
            response.raise_for_status()
            
            # 获取文件大小
            total_size = int(response.headers.get('content-length', 0))
            self.update_progress(f"文件大小: {total_size / 1024 / 1024:.2f} MB")
            
            # 写入文件
            with open(filepath, 'wb') as f:
                for chunk in response.iter_content(chunk_size=8192):
                    if chunk:
                        f.write(chunk)
            
            self.update_progress(f"文件已保存到: {filepath}")
            return True
            
        except requests.exceptions.RequestException as e:
            self.update_progress(f"下载文件时发生网络错误: {str(e)}")
            return False
        except Exception as e:
            self.update_progress(f"下载文件时发生错误: {str(e)}")
            return False

    def crawl_reports(self, years=None, selected_types=None, stock_name=None):
        """
        爬取指定年份和报告类型的财务报告
        
        Args:
            years: 要爬取的年份列表，如果为None则爬取近三年的报告
            selected_types: 选择的报告类型列表，如果为None则爬取所有类型的报告
            stock_name: 股票名称，用于创建文件夹
            
        Returns:
            str: 生成的Excel报告文件路径，如果没有找到报告则返回None
        """
        # 创建基础保存目录
        base_dir = "financial_reports"
        if not os.path.exists(base_dir):
            os.makedirs(base_dir)
            
        # 创建本次任务的专属目录
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        if stock_name:
            task_dir = os.path.join(base_dir, f"{stock_name}_{self.stock_code}_{timestamp}")
        else:
            task_dir = os.path.join(base_dir, f"{self.stock_code}_{timestamp}")
        os.makedirs(task_dir)
            
        # 如果未指定年份，默认爬取近三年的报告
        if years is None:
            current_year = datetime.now().year
            years = list(range(current_year - 2, current_year + 1))
            
        if years:
            years.sort()  # 确保年份是有序的
            start_date = datetime(years[0], 1, 1)
            end_date = datetime(years[-1], 12, 31)
        else:
            start_date = datetime(datetime.now().year - 2, 1, 1)
            end_date = datetime(datetime.now().year, 12, 31)
            
        # 如果未指定报告类型，默认爬取所有类型的报告
        if selected_types is None:
            selected_types = [t.report_name for t in ReportType]
            
        reports_data = []
        for type_name in selected_types:
            report_type = ReportType.from_name(type_name)
            if report_type is None:
                self.update_progress(f"未知的报告类型: {type_name}")
                continue
                
            report_list = self.get_report_list(start_date, end_date, report_type)
            pattern = {
                ReportType.ANNUAL: ["年度报告", "年报"],
                ReportType.SEMI_ANNUAL: ["半年度报告", "半年报"],
                ReportType.Q1: ["第一季度报告", "一季报", "第1季度报告", "2024年第一季度业绩预告"],
                ReportType.Q3: ["第三季度报告", "三季报", "第3季度报告"],
                ReportType.SOCIAL: ["社会责任报告", "企业社会责任报告"],
                ReportType.INTERNAL: ["内部控制报告", "内控报告", "内控自我评价报告"],
                ReportType.ESG: ["ESG报告", "环境、社会及管治报告", "环境社会治理报告"],
                ReportType.SUSTAINABLE: ["可持续发展报告"],
                ReportType.IPO_PROSPECTUS: ["招股说明书", "招股意向书", "发行保荐书"],
                ReportType.IPO_INQUIRY: ["问询函", "回复", "审核问询函"]
            }.get(report_type, [])
            
            for report in report_list:
                try:
                    title = report['title']
                    notice_date = report['notice_date']
                    date = datetime.strptime(notice_date, '%Y-%m-%d %H:%M:%S')
                    
                    # 检查年份
                    if years and date.year not in years:
                        continue
                        
                    # 检查报告类型
                    if not any(keyword in title for keyword in pattern):
                        continue
                        
                    # 下载PDF文件
                    download_url = f"https://pdf.dfcfw.com/pdf/H2_{report['art_code']}_1.pdf"
                    
                    # 添加随机延时
                    time.sleep(random.uniform(1, 2))
                    
                    # 下载PDF文件
                    pdf_response = requests.get(download_url, headers=self.headers)
                    if pdf_response.status_code == 200:
                        filename = os.path.join(task_dir, f"{title}_{date.strftime('%Y%m%d')}.pdf")
                        with open(filename, 'wb') as f:
                            f.write(pdf_response.content)
                        self.update_progress(f"已下载: {filename}")
                        
                        reports_data.append({
                            '序号': len(reports_data) + 1,
                            '文件名': os.path.basename(filename),
                            '发布日期': date.strftime('%Y-%m-%d'),
                            '报告类型': type_name,
                            '下载链接': download_url
                        })
                        
                except (ValueError, TypeError) as e:
                    self.update_progress(f"处理日期时出错 ({notice_date}): {str(e)}")
                    continue
                    
        # 生成Excel报告
        if reports_data:
            excel_file = os.path.join(task_dir, f"报告清单_{self.stock_code}_{timestamp}.xlsx")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "报告清单"
            
            # 设置表头
            headers = ['序号', '文件名', '发布日期', '报告类型', '下载链接']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = styles.Font(bold=True)
                cell.alignment = styles.Alignment(horizontal='center', vertical='center')
                
            # 写入数据
            for row, data in enumerate(reports_data, 2):
                for col, key in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col)
                    cell.value = data[key]
                    cell.alignment = styles.Alignment(horizontal='left', vertical='center', wrap_text=True)
                    
            # 调整列宽
            for col in ws.columns:
                max_length = 0
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[col[0].column_letter].width = min(adjusted_width, 50)
                
            wb.save(excel_file)
            return excel_file
            
        return None

if __name__ == "__main__":
    # 测试代码
    crawler = StockCrawler("300903")
    crawler.crawl_reports()
